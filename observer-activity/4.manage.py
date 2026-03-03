#!/usr/bin/env python3
"""
Management script for distributing images from multiple source directories among annotators
and creating tracking spreadsheets for manual classification with chronological renaming
"""

import os
import shutil
import argparse
import glob
import math
import random
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image

# List of annotators based on the image provided
ANNOTATORS = [
    "1-Julien",
    "2-Ahmad", 
    "3-Basel",
    "4-Rishi",
    "5-Sarah (Ehsani)",
    "6-Sara",
    "7-Mostafa",
    "8-Faeze"
]

class FileInfo:
    """Class to store information about files being processed."""
    def __init__(self, original_path, new_filename):
        self.original_path = original_path
        self.new_filename = new_filename
        self.original_filename = os.path.basename(original_path)
        self.source_directory = os.path.basename(os.path.dirname(original_path))

def resize_image_if_needed(image_path, target_size=(128, 128)):
    """
    Check if image is already target size, if not resize it
    
    Args:
        image_path (str): Path to the image file
        target_size (tuple): Target size (width, height)
    
    Returns:
        PIL.Image: Resized image object
    """
    try:
        with Image.open(image_path) as img:
            # Convert to RGB if necessary (handles RGBA, grayscale, etc.)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Check if already the target size
            if img.size == target_size:
                return img.copy()
            else:
                # Resize to target size
                resized_img = img.resize(target_size, Image.LANCZOS)
                return resized_img
    except Exception as e:
        print(f"Error processing image {image_path}: {e}")
        raise

def count_images_in_directories(directories):
    """Count total number of image files in multiple directories"""
    total_count = 0
    all_image_files = []
    
    for directory in directories:
        if not os.path.exists(directory):
            print(f"Warning: Directory {directory} does not exist!")
            continue
            
        # Find all image files in the directory
        image_files = []
        for ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
            pattern = os.path.join(directory, f"*{ext}")
            image_files.extend(glob.glob(pattern))
            pattern = os.path.join(directory, f"*{ext.upper()}")
            image_files.extend(glob.glob(pattern))
        
        print(f"  {directory}: {len(image_files)} images found")
        total_count += len(image_files)
        all_image_files.extend(image_files)
    
    # Shuffle the images to ensure random distribution among annotators
    random.shuffle(all_image_files)
    print(f"Total images found: {total_count} (shuffled for random distribution)")
    return total_count, all_image_files

def distribute_images_with_renaming(image_files, annotators):
    """
    Distribute images equally among annotators with chronological renaming
    
    Args:
        image_files (list): List of image file paths
        annotators (list): List of annotator names
    
    Returns:
        tuple: (distribution dict, original_filenames dict)
    """
    total_images = len(image_files)
    num_annotators = len(annotators)
    
    # Calculate images per annotator
    base_count = total_images // num_annotators
    remainder = total_images % num_annotators
    
    print(f"Total images: {total_images}")
    print(f"Number of annotators: {num_annotators}")
    print(f"Base images per annotator: {base_count}")
    print(f"Extra images to distribute: {remainder}")
    
    # Create distribution with renaming
    distribution = {}
    original_filenames = {}
    start_idx = 0
    
    for i, annotator in enumerate(annotators):
        # Some annotators get one extra image if there's a remainder
        count = base_count + (1 if i < remainder else 0)
        end_idx = start_idx + count
        
        # Get files for this annotator
        annotator_files = image_files[start_idx:end_idx]
        
        # Create file info with new chronological names
        file_info_list = []
        for j, file_path in enumerate(annotator_files):
            # Generate new filename with zero-padded numbering starting from 0000001
            new_filename = f"{j+1:07d}.png"
            
            file_info = {
                'original_path': file_path,
                'original_filename': os.path.basename(file_path),
                'new_filename': new_filename
            }
            
            file_info_list.append(file_info)
            original_filenames[new_filename] = os.path.basename(file_path)
        
        distribution[annotator] = file_info_list
        print(f"{annotator}: {count} images (0000001.png to {count:07d}.png)")
        
        start_idx = end_idx
    
    return distribution, original_filenames

def create_annotator_directories(base_output_dir, distribution, original_filenames):
    """
    Create directories for each annotator and copy their assigned images with chronological renaming
    Images are resized to 128x128 if they are not already that size
    
    Args:
        base_output_dir (str): Base directory to create annotator folders
        distribution (dict): Dictionary mapping annotators to their files
        original_filenames (dict): Dictionary mapping new filenames to original filenames
    
    Returns:
        dict: Mapping of annotator to their directory path
    """
    os.makedirs(base_output_dir, exist_ok=True)
    annotator_dirs = {}
    
    print(f"\nCreating annotator directories in: {base_output_dir}")
    print("Checking and resizing images to 128x128 if needed...")
    
    total_resized = 0
    total_processed = 0
    
    for annotator, file_info_list in distribution.items():
        # Create annotator directory
        annotator_dir = os.path.join(base_output_dir, annotator)
        os.makedirs(annotator_dir, exist_ok=True)
        annotator_dirs[annotator] = annotator_dir
        
        print(f"\nProcessing {len(file_info_list)} images for {annotator}")
        
        # Process images for this annotator
        for i, file_info in enumerate(file_info_list):
            original_path = file_info['original_path']
            new_filename = file_info['new_filename']
            dest_path = os.path.join(annotator_dir, new_filename)
            
            try:
                # Resize image if needed and save
                resized_img = resize_image_if_needed(original_path, target_size=(128, 128))
                
                # Check if resizing was needed
                with Image.open(original_path) as original_img:
                    if original_img.size != (128, 128):
                        total_resized += 1
                        print(f"  Resized {os.path.basename(original_path)} from {original_img.size} to (128, 128) -> {new_filename}")
                    else:
                        print(f"  Copied {os.path.basename(original_path)} (already 128x128) -> {new_filename}")
                
                # Save the processed image
                resized_img.save(dest_path, 'PNG', quality=95)
                total_processed += 1
                
            except Exception as e:
                print(f"  ERROR processing {original_path}: {e}")
                # Fallback: copy original file if processing fails
                shutil.copy2(original_path, dest_path)
                total_processed += 1
    
    print(f"\nImage processing summary:")
    print(f"  Total images processed: {total_processed}")
    print(f"  Images resized to 128x128: {total_resized}")
    print(f"  Images already 128x128: {total_processed - total_resized}")
    
    return annotator_dirs

def create_classification_spreadsheet(distribution, original_filenames, output_path):
    """
    Create Excel spreadsheet with sheets for each annotator
    
    Args:
        distribution (dict): Dictionary mapping annotators to their file info
        original_filenames (dict): Dictionary mapping new filenames to original filenames
        output_path (str): Path to save the Excel file
    """
    print(f"\nCreating classification spreadsheet: {output_path}")
    
    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create summary sheet first
    summary_ws = wb.create_sheet("Summary")
    summary_data = []
    
    for annotator, file_info_list in distribution.items():
        # Create sheet for each annotator
        ws = wb.create_sheet(annotator)
        
        # Prepare data for this annotator
        data = []
        for file_info in file_info_list:
            data.append({
                'Filename': file_info['new_filename'],
                'Original_Filename': file_info['original_filename'],
                'Predicted_Class': ''  # Empty for manual entry
            })
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add to summary data
        summary_data.append({
            'Annotator': annotator,
            'Total_Images': len(file_info_list),
            'Completed': 0,
            'Remaining': len(file_info_list)
        })
    
    # Create summary sheet
    summary_df = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        summary_ws.append(r)
    
    # Style summary sheet header
    for cell in summary_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust summary sheet columns
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    print(f"Spreadsheet created with {len(distribution)} annotator sheets")

def create_instructions_file(output_dir):
    """Create instructions file for annotators"""
    instructions_path = os.path.join(output_dir, "INSTRUCTIONS.txt")
    
    instructions = """
DENTAL IMAGE CLASSIFICATION INSTRUCTIONS
=======================================

Dear Annotators,

You have been assigned a set of dental images for manual classification. 
These images were initially predicted as Class 1+ by our AI model, but require 
human verification and detailed classification.

TASK:
-----
1. Review each image in your assigned folder
2. Classify each image according to the dental classification system
3. Enter your classification in the corresponding Excel spreadsheet

CLASSIFICATION CATEGORIES:
-------------------------
- Class 0: [Add description]
- Class 1: [Add description] 
- Class 2: [Add description]
- Class 3: [Add description]
- Class 4: [Add description]
- Class 5: [Add description]
- Class 6: [Add description]

SPREADSHEET USAGE:
-----------------
1. Open the Excel file: 'classification_tracking.xlsx'
2. Go to your assigned sheet (named with your name)
3. For each filename, enter the correct class in the 'Predicted_Class' column
4. Save the file regularly

IMPORTANT NOTES:
---------------
- Please classify ALL images assigned to you
- If you're unsure about an image, mark it as 'UNCERTAIN' and add a note
- Contact the team lead if you encounter any issues
- Expected completion time: [Add timeline]

Thank you for your contribution to this important classification task!

Contact: [Add contact information]
"""
    
    with open(instructions_path, 'w') as f:
        f.write(instructions)
    
    print(f"Instructions file created: {instructions_path}")

def main():
    parser = argparse.ArgumentParser(description='Distribute images from multiple directories among annotators')
    parser.add_argument('--input_dirs', '-i', nargs='+', required=True,
                       help='Directories containing images (e.g., syn and syn2)')
    parser.add_argument('--output_dir', '-o', required=True,
                       help='Base directory to create annotator folders')
    parser.add_argument('--spreadsheet', '-s', 
                       help='Path to save Excel tracking spreadsheet (default: classification_tracking.xlsx)')
    parser.add_argument('--annotators', nargs='+',
                       help='Custom list of annotator names (default: predefined list)')
    
    args = parser.parse_args()
    
    # Use custom annotators if provided, otherwise use default
    annotators = args.annotators if args.annotators else ANNOTATORS
    
    # Set default spreadsheet path if not provided
    spreadsheet_path = args.spreadsheet if args.spreadsheet else os.path.join(args.output_dir, 'classification_tracking.xlsx')
    
    try:
        print("="*80)
        print("DENTAL IMAGE DISTRIBUTION AND TRACKING SYSTEM")
        print("="*80)
        
        # Count images in input directories
        print(f"\nAnalyzing input directories: {args.input_dirs}")
        total_count, image_files = count_images_in_directories(args.input_dirs)
        
        if total_count == 0:
            print("No image files found in the input directories!")
            return 1
        
        # Distribute images among annotators with renaming
        print(f"\nDistributing images among {len(annotators)} annotators:")
        distribution, original_filenames = distribute_images_with_renaming(image_files, annotators)
        
        # Create annotator directories and copy images
        print(f"\nSetting up annotator workspaces...")
        annotator_dirs = create_annotator_directories(args.output_dir, distribution, original_filenames)
        
        # Create tracking spreadsheet
        print(f"\nCreating tracking spreadsheet...")
        create_classification_spreadsheet(distribution, original_filenames, spreadsheet_path)
        
        # Create instructions file
        create_instructions_file(args.output_dir)
        
        # Print summary
        print("\n" + "="*80)
        print("DISTRIBUTION COMPLETE!")
        print("="*80)
        print(f"Total images distributed: {total_count}")
        print(f"Number of annotators: {len(annotators)}")
        print(f"Source directories: {args.input_dirs}")
        print(f"Output directory: {args.output_dir}")
        print(f"Tracking spreadsheet: {spreadsheet_path}")
        
        print(f"\nAnnotator assignments:")
        for annotator, file_info_list in distribution.items():
            print(f"  {annotator}: {len(file_info_list)} images (0000001.png to {len(file_info_list):07d}.png)")
        
        print(f"\nNext steps:")
        print(f"1. Share the annotator folders with respective team members")
        print(f"2. Distribute the Excel tracking spreadsheet")
        print(f"3. Provide classification guidelines and deadlines")
        print(f"4. Monitor progress through the tracking spreadsheet")
        print(f"5. Files are renamed chronologically starting from 001.png for each annotator")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())
