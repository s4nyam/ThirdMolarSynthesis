#!/usr/bin/env python3
"""
Process annotated images by renaming them with their predicted class prefix.
For each user in the Excel file, rename files in their directory to include the predicted class.
Example: 0000001.png -> 0_0000001.png (if predicted class is 0)
"""

import openpyxl
import os
import shutil
from pathlib import Path
from difflib import SequenceMatcher

def match_user_directory(sheet_name, base_dir):
    """
    Match sheet name to directory name using longest string matching.
    Handles cases like "1-Julien" matching to "1-Julien" directly, 
    or "5-Sarah (Ehsani)" to the same directory name.
    """
    user_dirs = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d))]
    
    # Try exact match first
    if sheet_name in user_dirs:
        return sheet_name
    
    # Try longest string matching
    best_match = None
    best_ratio = 0
    
    for user_dir in user_dirs:
        ratio = SequenceMatcher(None, sheet_name.lower(), user_dir.lower()).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = user_dir
    
    return best_match if best_ratio > 0.6 else None


def process_user_files(sheet_name, sheet, user_dir_path, dry_run=False):
    """
    Process files for a user:
    - Read filenames and predicted classes from Excel sheet
    - Rename files to include predicted class prefix
    """
    if not os.path.exists(user_dir_path):
        print(f"  ⚠ Directory not found: {user_dir_path}")
        return 0, 0
    
    processed_count = 0
    skipped_count = 0
    
    # Iterate through rows (skip header)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        filename = row[0]
        predicted_class = row[1]
        
        if filename is None or predicted_class is None:
            continue
        
        # Convert to string and strip whitespace
        filename = str(filename).strip()
        predicted_class = str(predicted_class).strip()
        
        # Skip if filename is empty
        if not filename:
            continue
        
        # Create new filename with prefix
        new_filename = f"{predicted_class}_{filename}"
        
        old_path = os.path.join(user_dir_path, filename)
        new_path = os.path.join(user_dir_path, new_filename)
        
        # Check if old file exists
        if not os.path.exists(old_path):
            print(f"  ⚠ File not found: {filename}")
            skipped_count += 1
            continue
        
        # Check if new file already exists
        if os.path.exists(new_path):
            print(f"  ⚠ Target file already exists: {new_filename} (skipping)")
            skipped_count += 1
            continue
        
        # Rename the file
        if not dry_run:
            shutil.move(old_path, new_path)
        
        print(f"  ✓ {filename} → {new_filename}")
        processed_count += 1
    
    return processed_count, skipped_count


def main():
    excel_path = "/home/sj/working_dir/3m-project-128/3m-tg-dental-activity/annotator_workspace-provided/annotatot-provided-classification_tracking.xlsx"
    base_dir = "/home/sj/working_dir/3m-project-128/3m-tg-dental-activity/annotator_workspace-provided"
    
    # Check if Excel file exists
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Skip Summary sheet
    sheet_names = [name for name in wb.sheetnames if name != "Summary"]
    
    print(f"\nProcessing {len(sheet_names)} annotator sheets...\n")
    
    total_processed = 0
    total_skipped = 0
    
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        
        # Match directory to sheet name
        matched_dir = match_user_directory(sheet_name, base_dir)
        
        if matched_dir is None:
            print(f"✗ {sheet_name}: No matching directory found")
            continue
        
        user_dir_path = os.path.join(base_dir, matched_dir)
        
        print(f"Processing: {sheet_name}")
        print(f"  Directory: {matched_dir}")
        
        processed, skipped = process_user_files(sheet_name, sheet, user_dir_path, dry_run=False)
        
        total_processed += processed
        total_skipped += skipped
        
        if processed == 0 and skipped == 0:
            print(f"  (No files to process)")
        else:
            print(f"  Summary: {processed} renamed, {skipped} skipped")
        print()
    
    print(f"\n{'='*60}")
    print(f"Total: {total_processed} files renamed, {total_skipped} skipped")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
